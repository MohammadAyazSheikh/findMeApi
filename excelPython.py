import os.path
import openpyxl
from openpyxl import Workbook
wb = Workbook()
ws = wb.active


def createFile(path):

    ws['A1'] = 'hour'
    ws['B1'] = 'city'

    i = 1

    for x in range(0, 24):

        ws['A'+str(i)] = x
        ws['B'+str(i)] = 'no enough data'
        i = i+1

    wb.save(path)

# -------------------------------------


def insertValues(uID, city, hour, path):

    srcfile = openpyxl.load_workbook(path, read_only=False, keep_vba=True)
    ws = srcfile.active

    i = 0
    for x in range(0, 24):
        i = i+1
        hr = ws['A'+str(i)].value
        hrVal = str(hr)

        if hrVal == hour:
            print(hrVal)
            ct = ws['B'+str(i)].value

            if(ct == None):
                ws['B'+str(i)].value = city
            else:
                ct = str(ct)+','+city
                ws['B'+str(i)].value = ct

            srcfile.save(path)
            # wb.save(path)

# ------------------------------------


def addValues(uID, city, hour):

    path = 'user '+str(uID)+' file.xlsx'
    isExist = os.path.isfile(path)

    if isExist:
        # insertValues(uID,city,hour,path)
        insVal(path, int(hour), city)
    # else:
        # createFile(path)
        # insertValues(uID,city,hour,path)
        # insVal(path, int(hour), city)



# -------------------------------------------


def insVal(path, hr_, place):

    srcfile = openpyxl.load_workbook(path, read_only=False, keep_vba=True)
    ws = srcfile.active

    cityList = []
    hrList = []
    for x in range(2, 26):

        hr = ws['A'+str(x)].value
        city = ws['B'+str(x)].value

        cityList.append(str(city))
        hrList.append(str(hr))

    print(cityList, hrList)

    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'hour'
    ws['B1'] = 'city'

    i = 0
    for x in range(2, 26):  # saving old values
        ws['A'+str(x)] = i  # hrList[i]
        if x == hr_+2:
            ws['B'+str(x)] = str(cityList[i]+','+place)
            # print(place, hr_)
        else:
            ws['B'+str(x)] = cityList[i]
        i = i+1

    wb.save(path)


# --------------------------------------------------------
def addInitialData(uID, data):

    path = 'user '+str(uID)+' file.xlsx'
  
    ws['A1'] = 'hour'
    ws['B1'] = 'city'

    i = 2

    for x in range(0, 24):

        ws['A'+str(i)] = x
        ws['B'+str(i)] = data[x]
        i = i+1

    wb.save(path)


Data = ['test', 'test', 'test', 'test', 'test', 'test', 'test', 'test', 'test', 'test',
        'test', 'test', 'test', 'test', 'test', 'test',
        'test', 'test', 'test', 'test', 'test', 'test', 'test', 'test']


# addInitialData('Test_file1.xlsx',Data)

# insVal('file.xlsx',3,'hussainabad')
# createFile('file.xlsx')
# insVal('file.xlsx',3,'asa')
# addValues('xx','ara',3)
